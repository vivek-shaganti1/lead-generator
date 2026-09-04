"""
Processes the 175 Deep Research Intent Leads:
1. Ingests raw TSV from data/raw_new_leads.tsv.
2. Pre-flight checks: skips known bounced addresses and unresolvable domains.
3. Dispatches personalized AI agent & automation proposals via Google SSL SMTP (ksvdevlopers@gmail.com).
4. Tracks Gmail Sent, Instagram Sent, and Facebook Sent columns.
5. Saves 175_DEEP_RESEARCH_OMNICHANNEL_TRACKER.xlsx & CSV.
6. Synchronizes CRM Database & Master CRM Excel.
"""
from __future__ import annotations

import csv
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import os
from pathlib import Path
import smtplib
import sys
import time
import dns.resolver
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT / "backend"))

from app.config import settings
from app.db import SessionLocal, init_db
from app.models import (
    Business,
    Campaign,
    Deal,
    DealStage,
    EmailMessage,
    Lead,
    LeadStatus,
    MessageStatus,
)
from app.services.crm.excel_sync import trigger_master_excel_sync
from app.utils import new_token, utcnow

GMAIL_USER = settings.smtp_user
GMAIL_PASSWORD = settings.smtp_password
SENDER_NAME = "KSV AI & Automation Solutions"

KNOWN_BOUNCES = {
    'david.patel@micarehealth.com', 'sjohnson@platinumderm.com', 'jennifer.williams@texasoncology.com',
    'sjohnson@wholefamilyhealth.org', 'amanda.rodriguez@carbonhealth.com', 'nwalker@bomane.com',
    'brian.hall@purespaandsalon.com', 'james.wilson@universalmusic.com', 'mstevens@tiktok.com',
    'lisa.brown@andiamo-group.com', 'athompson@solve.io', 'brian.hall@whyhydrate.com',
    'mhernandez@finleysbarbershop.com', 'lauren.allen@salonvivace.com', 'lbrown@sollishealth.com',
    'lbrown@zetaglobal.com', 'rachel.green@colgatepalmolive.com', 'rachel.green@bpg-usa.com',
    'lisa.brown@iceye.com', 'alex.thompson@nissinfoods.com', 'lisa.brown@stepstonegroup.com',
    'mark.stevens@paretohealth.com', 'rgreen@pandora.net', 'james.wilson@uniqlo.com',
    'mark.stevens@yami.com', 'rachel.green@paretohealth.com', 'james.wilson@nift.com',
    'mark.stevens@continental.com', 'lisa.brown@mapei.com', 'mstevens@sluhn.org',
    'james.wilson@transcom.com', 'jwilson@bayclubs.com', 'vikram.singh@openai.com',
    'ewhite@zetaglobal.com', 'james.wilson@chipotle.com', 'psharma@ontra.ai',
    'mark.stevens@finalsite.com', 'sarah.johnson@archwellhealth.com', 'rachel.green@ernestpackaging.com',
    'ewhite@catena.com', 'ryan.clark@plentyone.com', 'stephanie.martinez@kw.com',
    'jessica.miller@enhancedlabs.com', 'rachel.green@scale.com', 'mstevens@column.com',
    'mstevens@inflect.com', 'mstevens@amazon.com', 'mstevens@notion.com',
    'mstevens@belk.com', 'mark.stevens@sweetgreen.com', 'chris.lee@vts.com',
    'chris.lee@firstam.com', 'jwilson@accenture.com', 'ankit.patel@intuit.com',
    'daniel.garcia@realtor.com', 'michael.chen@goldenstatedermatology.com'
}

UNRESOLVABLE_DOMAINS = {'klavae.com', 'wildsidenails.com', 'example.com', 'exampleclinic.com'}

def generate_consultative_proposal(lead: dict) -> tuple[str, str, str]:
    comp = lead.get('company', 'your team')
    contact = lead.get('dm_name') or 'Hiring Team'
    first_name = contact.split()[0] if contact else 'there'
    role = lead.get('job_title', 'Front Desk / Operations Specialist')
    niche = lead.get('niche', 'Operations')
    use_case = lead.get('use_case', 'AI Receptionist & Workflow Automation')
    location = lead.get('location', 'your market')

    subject = f"AI & Automation for {comp}'s {role} opening"

    body_text = f"""Hi {first_name},

I noticed that {comp} is actively recruiting for a {role} in {location}. 

Given the high operational volume typical in {niche}, hiring, training, and retaining full-time staff for front-desk coordination, inbound patient/customer communications, and appointment scheduling can easily consume $45,000–$65,000 annually per position—not including software overhead and turnover costs.

At KSV AI & Automation Solutions, we design and deploy autonomous AI agents and automated operational workflows built specifically for {niche} enterprises. Our systems:
• Handle 100% of routine inbound inquiries, scheduling, and intake instantly 24/7 across voice, email, and web.
• Integrate directly with your existing CRM, PMS/EHR, or scheduling platforms with zero disruption.
• Deliver human-level responsiveness and empathy while slashing administrative overhead by up to 80%.

Instead of a lengthy commitment, we would love to build a custom, interactive prototype tailored to {comp}'s workflow completely free of charge. If it saves your team significant hours in the first 14 days, we can discuss an introductory pilot; if not, there is zero obligation.

Would you be open to a brief 7-minute introductory demo this week?

Best regards,

Vivek Shaganti
Founder & Lead Automation Engineer
KSV AI & Automation Solutions
Direct: shagantivivekgoud@gmail.com
Outreach: ksvdevlopers@gmail.com
"""

    body_html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; line-height: 1.6; color: #1e293b; background-color: #f8fafc; margin: 0; padding: 24px; }}
  .container {{ max-width: 620px; margin: 0 auto; background: #ffffff; border-radius: 12px; border: 1px solid #e2e8f0; overflow: hidden; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05); }}
  .header {{ background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 100%); padding: 32px 28px; color: #ffffff; }}
  .header h2 {{ margin: 0 0 8px 0; font-size: 22px; font-weight: 700; color: #ffffff; }}
  .header p {{ margin: 0; color: #93c5fd; font-size: 14px; letter-spacing: 0.5px; text-transform: uppercase; }}
  .content {{ padding: 28px; font-size: 15px; color: #334155; }}
  .highlight-card {{ background: #f0fdf4; border-left: 4px solid #16a34a; padding: 16px 20px; border-radius: 0 8px 8px 0; margin: 20px 0; }}
  .metric {{ font-size: 20px; font-weight: 700; color: #15803d; }}
  .features {{ margin: 20px 0; padding-left: 20px; }}
  .features li {{ margin-bottom: 10px; }}
  .cta-btn {{ display: inline-block; background: #2563eb; color: #ffffff !important; padding: 12px 26px; border-radius: 8px; font-weight: 600; text-decoration: none; margin-top: 16px; box-shadow: 0 2px 4px rgba(37,99,235,0.2); }}
  .footer {{ background: #f1f5f9; padding: 20px 28px; border-top: 1px solid #e2e8f0; font-size: 13px; color: #64748b; }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <p>Operational Intelligence & AI Automation</p>
    <h2>Solving Staffing & Intake Friction for {comp}</h2>
  </div>
  <div class="content">
    <p>Hi <strong>{first_name}</strong>,</p>
    <p>I noticed that <strong>{comp}</strong> is actively recruiting for a <strong>{role}</strong> in {location}.</p>
    <p>Given the operational workload typical in {niche}, filling and onboarding dedicated front-desk and support roles often costs upwards of <strong>$45,000–$65,000/year</strong> per headcount—all while teams still face after-hours delays and call bottlenecks.</p>
    <div class="highlight-card">
      <div class="metric">80% Overhead Reduction & 24/7 Coverage</div>
      <p style="margin: 6px 0 0 0; color: #166534; font-size: 14px;">Our custom AI Agents seamlessly handle routine inquiries, triage, scheduling, and intake with zero hold times.</p>
    </div>
    <p>Key Capabilities Engineered for {comp}:</p>
    <ul class="features">
      <li><strong>Autonomous Communication:</strong> Human-grade conversational triage via email, web chat, and voice.</li>
      <li><strong>Frictionless Integration:</strong> Syncs bi-directionally with your scheduling and CRM stack.</li>
      <li><strong>Immediate Availability:</strong> 24/7 response time with 100% data capture and zero missed prospects.</li>
    </ul>
    <p>We would be thrilled to build an interactive, working prototype tailored specifically to {comp}'s {role} workflow <strong>at zero cost or commitment</strong>. If it demonstrates clear ROI, we can discuss an introductory pilot; if not, there is no obligation whatsoever.</p>
    <p style="text-align: center; margin: 28px 0 16px 0;">
      <a href="mailto:ksvdevlopers@gmail.com?subject=Re:%20AI%20Automation%20Prototype%20for%20{comp}" class="cta-btn">Review 7-Minute Interactive Demo</a>
    </p>
  </div>
  <div class="footer">
    <strong>KSV AI & Automation Solutions</strong> &bull; Vivek Shaganti<br>
    Enterprise Automation & Intelligent Agent Infrastructure<br>
    Hyderabad, India &bull; Direct: <a href="mailto:ksvdevlopers@gmail.com" style="color: #2563eb;">ksvdevlopers@gmail.com</a><br>
    <span style="font-size: 11px; color: #94a3b8;">If you prefer not to receive automation research, reply "Unsubscribe" and your email will be permanently removed.</span>
  </div>
</div>
</body>
</html>"""
    return subject, body_text, body_html

def main():
    print("=" * 60)
    print("INGESTING 175 DEEP RESEARCH INTENT LEADS")
    print("=" * 60)

    raw_file = Path("data/raw_new_leads.tsv")
    if not raw_file.exists():
        print("Error: data/raw_new_leads.tsv not found!")
        return

    lines = raw_file.read_text(encoding="utf-8").splitlines()
    raw_leads = []
    for line in lines:
        p = line.strip().split("\t")
        if len(p) > 23 and p[1].startswith("LEAD-"):
            raw_leads.append({
                'lead_id': p[1],
                'niche': p[2],
                'job_title': p[3],
                'job_board': p[5],
                'posting_date': p[6],
                'location': p[7],
                'company': p[8],
                'company_ig': p[11],
                'company_fb': p[12],
                'dm_ig': p[13],
                'dm_fb': p[14],
                'industry': p[15],
                'company_size': p[16],
                'company_hq': p[17],
                'dm_name': p[18],
                'dm_title': p[19],
                'email': p[21].strip().lower(),
                'status': p[22],
                'domain': p[23].strip().lower(),
                'intent_score': p[24],
                'intent_reason': p[25],
                'use_case': p[26],
                'priority': p[27],
                'research_date': p[28],
                'notes': p[29] if len(p) > 29 else ''
            })

    print(f"Total leads parsed: {len(raw_leads)}")

    # Connect SMTP
    print(f"Connecting to Google SMTP as {GMAIL_USER}...")
    server = smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15)
    server.login(GMAIL_USER, GMAIL_PASSWORD)
    print("SMTP connection established successfully!")

    # Connect DB
    init_db()
    db = SessionLocal()
    campaign = db.query(Campaign).first()
    if not campaign:
        campaign = Campaign(
            name="Deep Research Intent AI Campaign",
            subject_template="",
            body_template="",
            is_active=True
        )
        db.add(campaign)
        db.commit()
        db.refresh(campaign)

    sent_emails = set()
    results = []
    sent_count = 0
    skipped_bounce_count = 0
    skipped_domain_count = 0

    for idx, lead in enumerate(raw_leads, 1):
        lead_id = lead['lead_id']
        email_addr = lead['email']
        domain = lead['domain']
        comp = lead['company']

        gmail_status = ""
        ig_status = f"⏳ Queued ({lead['company_ig']})" if lead['company_ig'] != "[object Object]" else "⏳ Queued (Awaiting Handle)"
        fb_status = "⏳ Queued (Awaiting Platform Login)"

        # Check domain validity
        if domain in UNRESOLVABLE_DOMAINS:
            gmail_status = "❌ Skipped (No MX Server / Invalid Domain)"
            skipped_domain_count += 1
            results.append((lead, gmail_status, ig_status, fb_status))
            continue

        # Check known bounces
        if email_addr in KNOWN_BOUNCES:
            gmail_status = "❌ Skipped (Known Bounced Address - Address Not Found)"
            skipped_bounce_count += 1
            results.append((lead, gmail_status, ig_status, fb_status))
            continue

        # Check self-sending
        if email_addr == GMAIL_USER.lower():
            gmail_status = "❌ Skipped (Self-sending protection)"
            results.append((lead, gmail_status, ig_status, fb_status))
            continue

        # Deduplication: if email already sent in this batch
        if email_addr in sent_emails:
            gmail_status = "✔ Sent (Covered in Batch Dispatch)"
            results.append((lead, gmail_status, ig_status, fb_status))
            continue

        # Generate proposal
        subj, body_txt, body_html = generate_consultative_proposal(lead)

        msg = MIMEMultipart("alternative")
        msg["From"] = f"{SENDER_NAME} <{GMAIL_USER}>"
        msg["To"] = email_addr
        msg["Subject"] = subj
        msg["Reply-To"] = GMAIL_USER

        msg.attach(MIMEText(body_txt, "plain", "utf-8"))
        msg.attach(MIMEText(body_html, "html", "utf-8"))

        try:
            server.sendmail(GMAIL_USER, [email_addr], msg.as_string())
            sent_emails.add(email_addr)
            sent_count += 1
            gmail_status = "✔ Sent (Direct Dispatch)"
            print(f"[{sent_count:3d}] DISPATCHED -> {email_addr} ({comp} - {lead['job_title']})")
            time.sleep(0.4) # smooth rate limit

            # Update DB
            bus = db.query(Business).filter(Business.name == comp).first()
            if not bus:
                bus = Business(
                    name=comp,
                    website=f"https://{domain}",
                    category=lead['niche'],
                    country="US",
                    city=lead['location'].split(",")[0].strip() if "," in lead['location'] else lead['location'],
                    is_active=True
                )
                db.add(bus)
                db.flush()

            db_lead = db.query(Lead).filter(Lead.email == email_addr).first()
            if not db_lead:
                db_lead = Lead(
                    business_id=bus.id,
                    first_name=lead['dm_name'].split()[0] if lead['dm_name'] else "Team",
                    last_name=" ".join(lead['dm_name'].split()[1:]) if lead['dm_name'] and " " in lead['dm_name'] else "",
                    email=email_addr,
                    job_title=lead['dm_title'],
                    status=LeadStatus.CONTACTED,
                    lead_score=int(lead['intent_score']) if lead['intent_score'].isdigit() else 85
                )
                db.add(db_lead)
                db.flush()
            else:
                db_lead.status = LeadStatus.CONTACTED

            deal = db.query(Deal).filter(Deal.business_id == bus.id).first()
            if not deal:
                deal = Deal(
                    business_id=bus.id,
                    campaign_id=campaign.id,
                    title=f"{comp} - AI Agent & Workflow Automation",
                    stage=DealStage.OUTREACH_SENT,
                    deal_value=1200.0,
                    probability=0.20,
                    expected_revenue=240.0
                )
                db.add(deal)
                db.flush()

            email_rec = EmailMessage(
                lead_id=db_lead.id,
                deal_id=deal.id,
                direction="OUTBOUND",
                subject=subj,
                body=body_txt,
                status=MessageStatus.SENT,
                sent_at=utcnow()
            )
            db.add(email_rec)

        except Exception as ex:
            print(f"Error dispatching to {email_addr}: {ex}")
            gmail_status = f"❌ Error ({str(ex)[:40]})"

        results.append((lead, gmail_status, ig_status, fb_status))

    server.quit()
    db.commit()

    print("\n" + "=" * 60)
    print(f"DISPATCH SUMMARY: {sent_count} Dispatched, {skipped_bounce_count} Skipped Bounces, {skipped_domain_count} Skipped Domains")
    print("=" * 60)

    # 5. Build Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Omnichannel Intent Leads"
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Lead ID", "Niche", "Job Title", "Job Board", "Posting Date", "Location",
        "Company Name", "Company Instagram", "Company Facebook",
        "DM Full Name", "DM Job Title", "DM Instagram", "DM Facebook",
        "Work Email", "Company Domain", "Intent Score", "Intent Reason",
        "Automation Use Case", "Priority", "Gmail Sent", "Instagram Sent", "Facebook Sent",
        "Research Date", "Deep Research Notes"
    ]
    ws.append(headers)

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    regular_font = Font(name="Segoe UI", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    green_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    amber_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align

    for lead, g_stat, ig_stat, fb_stat in results:
        row_vals = [
            lead['lead_id'], lead['niche'], lead['job_title'], lead['job_board'], lead['posting_date'], lead['location'],
            lead['company'], lead['company_ig'], lead['company_fb'],
            lead['dm_name'], lead['dm_title'], lead['dm_ig'], lead['dm_fb'],
            lead['email'], lead['domain'], lead['intent_score'], lead['intent_reason'],
            lead['use_case'], lead['priority'], g_stat, ig_stat, fb_stat,
            lead['research_date'], lead['notes']
        ]
        ws.append(row_vals)
        cur_row = ws.max_row
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws.cell(row=cur_row, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 4, 5, 16, 19]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            # Color highlight for platform status columns
            if col_idx == 20: # Gmail Sent
                if "✔" in g_stat:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            elif col_idx in [21, 22]: # IG & FB
                cell.fill = amber_fill

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    excel_path = Path("data/175_DEEP_RESEARCH_OMNICHANNEL_TRACKER.xlsx")
    wb.save(excel_path)
    print(f"Saved Excel tracker: {excel_path} ({excel_path.stat().st_size} bytes)")

    # Save CSV
    csv_path = Path("data/175_DEEP_RESEARCH_OMNICHANNEL_TRACKER.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    print(f"Saved CSV tracker: {csv_path} ({csv_path.stat().st_size} bytes)")

    # Trigger CRM Master Excel Sync
    trigger_master_excel_sync(db)
    db.close()
    print("Synchronized CRM Database and Master Excel!")

if __name__ == "__main__":
    main()
