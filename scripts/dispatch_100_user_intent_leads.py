"""
Processes the 100 User Intent Leads:
1. Pre-flight DNS MX and domain verification.
2. Generates personalized human-grade, high-converting AI agent & automation emails.
3. Dispatches directly to prospect work emails via Google SMTP.
4. Adds 3 tracking columns: Gmail Sent, Instagram Sent, Facebook Sent.
5. Updates CRM Database (Businesses, Leads, Deals, Messages).
6. Generates 100_INTENT_LEADS_OMNICHANNEL_TRACKER.xlsx, CSV, and updates Master CRM Excel.
"""
from __future__ import annotations

import csv
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
from pathlib import Path
import smtplib
import sys
import time
import dns.resolver
import dns.exception
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT / "backend"))

from app.config import settings
from app.db import SessionLocal, init_db
from app.models import (
    Business,
    Campaign,
    Deal,
    DealStage,
    EmailMessage,
    Lead,
    LeadStatus,
    MessageStatus,
)
from app.services.crm.excel_sync import trigger_master_excel_sync
from app.utils import new_token, utcnow

GMAIL_USER = os.getenv("SMTP_USER", getattr(settings, "smtp_user", "ksvdevlopers@gmail.com"))
GMAIL_PASSWORD = os.getenv("SMTP_PASSWORD", getattr(settings, "smtp_password", "kztzxmkbrwhhtdzd"))
SENDER_NAME = "KSV AI & Automation Solutions"

custom_resolver = dns.resolver.Resolver()
custom_resolver.timeout = 2.0
custom_resolver.lifetime = 2.5


def check_domain_mx(domain: str, cache: dict[str, bool]) -> bool:
    domain = (domain or "").strip().lower()
    if not domain or "." not in domain or domain in ["example.com", "exampleclinic.com", "stealth-d2c.com"]:
        return False
    if domain in cache:
        return cache[domain]
    try:
        answers = custom_resolver.resolve(domain, "MX")
        has_mx = len(answers) > 0
        cache[domain] = has_mx
        return has_mx
    except Exception:
        try:
            answers_a = custom_resolver.resolve(domain, "A")
            has_a = len(answers_a) > 0
            cache[domain] = has_a
            return has_a
        except Exception:
            cache[domain] = False
            return False


def build_email_content(lead_data: dict[str, str]) -> tuple[str, str, str]:
    first_name = lead_data.get("Decision Maker First Name", "").strip() or "Hiring Leader"
    company_name = lead_data.get("Company Name", "").strip() or "your company"
    job_title = lead_data.get("Job Title", "").strip() or "active role"
    location = lead_data.get("Job Location", "").strip() or "your area"
    use_case = lead_data.get("Automation Use Case", "").strip()
    niche = lead_data.get("Niche", "").strip()

    subject = f"AI & Automation for {company_name}'s {job_title} opening"

    # Break use cases into readable bullet points
    if use_case:
        bullets = [f"• {u.strip().capitalize()}" for u in use_case.split("/")[:3]]
        bullet_text = "\n".join(bullets)
        bullet_html = "".join([f"<li style='margin-bottom: 6px;'>{u.strip().capitalize()}</li>" for u in use_case.split("/")[:3]])
    else:
        bullet_text = "• 24/7 Inbound inquiry handling & triage\n• Automated appointment booking & scheduling\n• Instant CRM data synchronization"
        bullet_html = "<li style='margin-bottom: 6px;'>24/7 Inbound inquiry handling & triage</li><li style='margin-bottom: 6px;'>Automated appointment booking & scheduling</li><li style='margin-bottom: 6px;'>Instant CRM data synchronization</li>"

    plain_body = f"""Hi {first_name},

I noticed {company_name} is actively hiring for a {job_title} in {location}.

Given the high administrative workload and hiring turnaround for this position, I wanted to reach out directly. We engineer custom, production-ready AI Agents & Workflow Automations that handle these operational bottlenecks autonomously:

{bullet_text}
• Zero hold times with instant, human-like voice and chat interactions
• Seamless 2-way sync with your existing CRM and software stack

Key Benefits for {company_name}:
1. 80% Lower Cost: Operates 24/7 at a small fraction of the cost of a full-time hire.
2. Rapid Deployment: Live and fine-tuned to your exact protocols within days, not months.
3. Introductory Pilot Program: We are offering minimal setup charges and generous introductory discounts for qualified organizations this quarter.

Would you be open to a quick 10-minute visual walkthrough of a working prototype built for {company_name} this week?

Simply reply to this email, and I'll share the preview with you.

Best regards,

{SENDER_NAME}
Enterprise AI Agents & Workflow Automation
Email: {GMAIL_USER}
"""

    styled_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 24px 0; background-color: #f8f7f4; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; color: #1e293b;">
  <table width="100%" border="0" cellspacing="0" cellpadding="0" style="background-color: #f8f7f4;">
    <tr>
      <td align="center" style="padding: 12px 16px;">
        <table width="100%" border="0" cellspacing="0" cellpadding="0" style="max-width: 600px; background-color: #ffffff; border: 1px solid #e2e8f0; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.04);">
          <tr>
            <td height="4" style="background: linear-gradient(90deg, #2563eb, #38bdf8); font-size: 0; line-height: 0;">&nbsp;</td>
          </tr>
          <tr>
            <td style="padding: 32px 32px 28px 32px;">
              <div style="margin-bottom: 20px;">
                <span style="display: inline-block; padding: 4px 12px; background-color: #eff6ff; color: #1d4ed8; font-size: 12px; font-weight: 700; border-radius: 9999px; text-transform: uppercase; letter-spacing: 0.04em;">
                  {company_name}
                </span>
                <span style="display: inline-block; margin-left: 8px; padding: 4px 10px; background-color: #f1f5f9; color: #475569; font-size: 12px; font-weight: 600; border-radius: 9999px;">
                  {job_title}
                </span>
                <span style="display: inline-block; margin-left: 8px; padding: 4px 10px; background-color: #ecfdf5; color: #047857; font-size: 12px; font-weight: 600; border-radius: 9999px;">
                  {niche}
                </span>
              </div>

              <p style="margin: 0 0 16px 0; font-size: 16px; font-weight: 600; color: #0f172a;">
                Hi {first_name},
              </p>

              <p style="margin: 0 0 16px 0; font-size: 15px; line-height: 1.6; color: #334155;">
                I noticed that <strong>{company_name}</strong> is currently hiring for a <strong>{job_title}</strong> in {location}.
              </p>

              <p style="margin: 0 0 16px 0; font-size: 15px; line-height: 1.6; color: #334155;">
                Given the administrative workload and recruitment turnaround for this role, we help teams automate these exact workflows using <strong>custom Autonomous AI Agents</strong>:
              </p>

              <div style="margin: 20px 0; padding: 18px 20px; background-color: #f8fafc; border-left: 4px solid #2563eb; border-radius: 6px;">
                <div style="font-size: 14px; font-weight: 700; color: #0f172a; margin-bottom: 10px;">
                  ⚡ Key Operational Capabilities:
                </div>
                <ul style="margin: 0; padding-left: 20px; font-size: 14px; line-height: 1.5; color: #475569;">
                  {bullet_html}
                  <li style="margin-bottom: 6px;">Zero hold times with instant, human-like voice and chat responses</li>
                  <li>Seamless 2-way synchronization with your existing CRM and tech stack</li>
                </ul>
              </div>

              <div style="margin: 20px 0; padding: 16px 18px; background-color: #fefce8; border: 1px solid #fef08a; border-radius: 8px;">
                <strong style="color: #854d0e; font-size: 14px;">💡 Why leaders choose AI Agents over manual hires:</strong>
                <p style="margin: 6px 0 0 0; font-size: 13px; line-height: 1.5; color: #713f12;">
                  • <strong>80% Lower Cost:</strong> Operates 24/7 for a fraction of a full-time salary.<br>
                  • <strong>Fast Turnaround:</strong> Deployed and customized for {company_name} in days.<br>
                  • <strong>Introductory Discount:</strong> Minimal pilot setup fees for new enterprise partners this quarter.
                </p>
              </div>

              <p style="margin: 20px 0 24px 0; font-size: 15px; line-height: 1.6; color: #334155;">
                Would you be open to a quick 10-minute walkthrough of an interactive prototype prepared for <strong>{company_name}</strong> this week?
              </p>

              <table border="0" cellspacing="0" cellpadding="0" style="margin: 24px 0 16px 0;">
                <tr>
                  <td align="center" style="border-radius: 8px; background-color: #0f172a;">
                    <a href="mailto:{GMAIL_USER}?subject=Re:%20AI%20Demo%20for%20{company_name}" target="_blank" style="font-size: 14px; font-weight: 600; color: #ffffff; text-decoration: none; padding: 12px 24px; border-radius: 8px; display: inline-block;">
                      Schedule 10-Min Demo &rarr;
                    </a>
                  </td>
                </tr>
              </table>

              <div style="margin-top: 28px; padding-top: 20px; border-top: 1px solid #f1f5f9; font-size: 13px; line-height: 1.5; color: #64748b;">
                Best regards,<br>
                <strong style="color: #0f172a; font-size: 14px;">{SENDER_NAME}</strong><br>
                <span>Enterprise Autonomous Agents &amp; Workflow Automation</span><br>
                <span style="color: #2563eb;">{GMAIL_USER}</span>
              </div>
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""

    return subject, plain_body, styled_html


def process_all_100_leads():
    print("=" * 80, flush=True)
    print("🚀 PROCESSING 100 USER INTENT LEADS (VALIDATION, DISPATCH & CRM SYNC)", flush=True)
    print("=" * 80, flush=True)

    tsv_path = _REPO_ROOT / "data" / "raw_100_leads.tsv"
    if not tsv_path.exists():
        print(f"❌ Error: {tsv_path} not found!", flush=True)
        return

    with open(tsv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        rows = list(reader)

    print(f"📊 Loaded {len(rows)} leads from {tsv_path}", flush=True)

    init_db()
    db = SessionLocal()
    campaign = db.query(Campaign).filter(Campaign.name == "Default outreach").first()

    mx_cache: dict[str, bool] = {}
    enriched_records = []
    dispatched_count = 0
    skipped_count = 0
    now = utcnow()

    server = None
    try:
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=25)
        server.login(GMAIL_USER, GMAIL_PASSWORD)
        print("✔ Authenticated to Google SMTP SSL successfully.\n", flush=True)
    except Exception as e:
        print(f"❌ Warning: SMTP connection failed: {e}. Will proceed with verification & tracker.", flush=True)

    for idx, row in enumerate(rows, 1):
        lead_id = row.get("Lead ID", f"LEAD-{idx:03d}").strip()
        comp_name = row.get("Company Name", "").strip()
        work_email = row.get("Work Email", "").strip().lower()
        domain = row.get("Company Domain", "").strip().lower()
        job_title = row.get("Job Title", "").strip()
        contact_name = row.get("Decision Maker Full Name", "").strip() or row.get("Decision Maker First Name", "").strip()
        intent_score = float(row.get("Intent Score", "75").strip() or "75")
        use_case = row.get("Automation Use Case", "").strip()

        # Pre-flight MX check
        has_mx = check_domain_mx(domain, mx_cache)

        # Build record with 3 new tracking columns
        rec = dict(row)
        rec["Instagram Sent"] = "⏳ Queued (Awaiting Platform Login)"
        rec["Facebook Sent"] = "⏳ Queued (Awaiting Platform Login)"

        if not has_mx or work_email == GMAIL_USER.lower():
            rec["Gmail Sent"] = "❌ Skipped (No MX Server / Invalid Domain)"
            skipped_count += 1
            print(f"[{idx:03d}/100] ❌ SKIPPED {lead_id} | {comp_name:<28} | {work_email:<34} (No MX Server)", flush=True)
        else:
            # Dispatch Email
            subject, plain_body, styled_html = build_email_content(row)

            if server:
                try:
                    msg = MIMEMultipart("alternative")
                    msg["Subject"] = subject
                    msg["From"] = f"{SENDER_NAME} <{GMAIL_USER}>"
                    msg["To"] = work_email
                    msg["Reply-To"] = GMAIL_USER
                    msg["X-Lead-ID"] = lead_id

                    msg.attach(MIMEText(plain_body, "plain", "utf-8"))
                    msg.attach(MIMEText(styled_html, "html", "utf-8"))

                    server.sendmail(GMAIL_USER, [work_email], msg.as_string())
                    rec["Gmail Sent"] = "✔ Sent"
                    dispatched_count += 1
                    print(f"[{idx:03d}/100] 🚀 SENT {lead_id} | {comp_name:<28} | {work_email:<34} | MX: OK", flush=True)
                except Exception as ex:
                    rec["Gmail Sent"] = f"❌ Error: {ex}"
                    print(f"[{idx:03d}/100] ❌ SMTP FAIL {lead_id} | {comp_name}: {ex}", flush=True)
            else:
                rec["Gmail Sent"] = "✔ Validated (Ready to Send)"

            # Database Tracking
            biz = db.query(Business).filter(Business.name == comp_name).first()
            if not biz:
                biz = Business(
                    source="user_100_intent_leads",
                    source_id=f"intent_{idx}_{lead_id}",
                    dedupe_key=f"intent:{idx}:{comp_name.lower()}",
                    name=comp_name,
                    category=row.get("Niche", "Technology"),
                    email=work_email,
                    city=row.get("Job Location", "").split(",")[0].strip(),
                    country_code="US",
                )
                db.add(biz)
                db.flush()

            lead_obj = db.query(Lead).filter(Lead.email == work_email).first()
            if not lead_obj:
                lead_obj = Lead(
                    business_id=biz.id,
                    campaign_id=campaign.id if campaign else None,
                    email=work_email,
                    contact_name=contact_name,
                    status=LeadStatus.CONTACTED if rec["Gmail Sent"] == "✔ Sent" else LeadStatus.NEW,
                    score=intent_score,
                    approved=True,
                    unsubscribe_token=new_token(32),
                    last_contacted_at=now,
                )
                db.add(lead_obj)
                db.flush()

            if rec["Gmail Sent"] == "✔ Sent":
                out_msg = EmailMessage(
                    lead_id=lead_obj.id,
                    step=0,
                    direction="out",
                    to_email=work_email,
                    from_email=GMAIL_USER,
                    subject=subject,
                    body_text=plain_body,
                    body_html=styled_html,
                    status=MessageStatus.SENT,
                    sent_at=now,
                    message_id=f"user-msg-{lead_id}-{new_token(8)}",
                )
                db.add(out_msg)

                deal = db.query(Deal).filter(Deal.lead_id == lead_obj.id).first()
                if not deal:
                    deal_val = 850.0 if "VERY HIGH" in row.get("Priority", "") else 650.0
                    deal = Deal(
                        lead_id=lead_obj.id,
                        business_id=biz.id,
                        title=f"AI Agent & Automation — {comp_name}",
                        company_name=comp_name,
                        contact_name=contact_name,
                        contact_email=work_email,
                        stage=DealStage.CONTACTED,
                        value=deal_val,
                        probability=30.0,
                        expected_close_at=now + datetime.timedelta(days=21),
                        notes=f"Actively hiring for {job_title}. Use case: {use_case}",
                    )
                    db.add(deal)

            if idx % 10 == 0:
                db.commit()

            time.sleep(0.35)

        enriched_records.append(rec)

    if server:
        try:
            server.quit()
        except Exception:
            pass

    db.commit()

    print("\n" + "=" * 80, flush=True)
    print(f"📊 CAMPAIGN DISPATCH COMPLETE!", flush=True)
    print(f"  • Total Leads Processed: {len(enriched_records)}", flush=True)
    print(f"  • Successfully Dispatched: {dispatched_count}", flush=True)
    print(f"  • Skipped (No MX / Dead Domain): {skipped_count}", flush=True)
    print("=" * 80, flush=True)

    # Save to dedicated Excel tracker with styling
    export_tracker_excel(enriched_records)
    export_tracker_csv(enriched_records)

    # Trigger 9-sheet Master CRM Excel update
    trigger_master_excel_sync(db)
    db.close()


def export_tracker_excel(records: list[dict[str, str]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "100_INTENT_LEADS_TRACKER"
    ws.views.sheetView[0].showGridLines = True

    if not records:
        return

    # Columns list: Original columns + 3 new columns
    orig_cols = list(records[0].keys())
    # Ensure Gmail Sent, Instagram Sent, Facebook Sent are clearly placed
    for c in ["Gmail Sent", "Instagram Sent", "Facebook Sent"]:
        if c in orig_cols:
            orig_cols.remove(c)
    headers = orig_cols + ["Gmail Sent", "Instagram Sent", "Facebook Sent"]

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    sent_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    sent_font = Font(name="Calibri", size=10, bold=True, color="047857")
    skip_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    skip_font = Font(name="Calibri", size=10, color="B91C1C")
    queued_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    queued_font = Font(name="Calibri", size=10, color="B45309")

    for row_idx, rec in enumerate(records, 2):
        row_vals = [rec.get(h, "") for h in headers]
        ws.append(row_vals)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = align_left

            val = str(cell.value)
            if "✔ Sent" in val:
                cell.fill = sent_fill
                cell.font = sent_font
                cell.alignment = align_center
            elif "❌" in val:
                cell.fill = skip_fill
                cell.font = skip_font
            elif "⏳" in val:
                cell.fill = queued_fill
                cell.font = queued_font
                cell.alignment = align_center

    # Column widths auto-adjust
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col[:40])
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    out_path = _REPO_ROOT / "data" / "100_INTENT_LEADS_OMNICHANNEL_TRACKER.xlsx"
    wb.save(out_path)
    print(f"✔ Saved Omnichannel Tracker Excel: {out_path} ({out_path.stat().st_size:,} bytes)", flush=True)


def export_tracker_csv(records: list[dict[str, str]]):
    if not records:
        return
    headers = list(records[0].keys())
    out_path = _REPO_ROOT / "data" / "100_INTENT_LEADS_OMNICHANNEL_TRACKER.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(records)
    print(f"✔ Saved Omnichannel Tracker CSV:   {out_path} ({out_path.stat().st_size:,} bytes)", flush=True)


if __name__ == "__main__":
    process_all_100_leads()
